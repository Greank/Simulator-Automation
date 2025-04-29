from openpyxl import load_workbook, Workbook
from datetime import datetime
import os

def write_custom_excel(results, index, device_id, filename="ur_stats.xlsx"):
    from openpyxl import Workbook, load_workbook
    from datetime import datetime
    import os

    ur_keys = [f"UR_{i:02}" for i in range(1, 15)]  # UR_01 ~ UR_14
    full_filename = f"{device_id}_{filename}"

    # 建立新 Excel（如果不存在）
    if not os.path.exists(full_filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "統計紀錄"
        headers = ["編號"] + ur_keys + ["total", "時間"]
        ws.append(headers)
        wb.save(full_filename)

    # 載入 Excel
    wb = load_workbook(full_filename)
    ws = wb.active

    row_data = []
    total = 0

    for key in ur_keys:
        count = 0
        for result_key, value in results.items():
            # 把所有以 UR_01 開頭的 key 加總起來
            if result_key.startswith(key):
                count += value
        row_data.append(count)
        total += count

    now = datetime.now().strftime("%Y/%m/%d %H:%M")
    final_row = [index] + row_data + [total, now]
    ws.append(final_row)

    wb.save(full_filename)
    print(f"✅ 編號 {index} 的統計已寫入 {full_filename}")

def load_config_to_device_map(device_map, filename="config.txt"):
    with open(filename, "r") as f:
        for line in f:
            if "=" in line:
                key, value = line.strip().split("=")
                if key in device_map:
                    device_map[key][0] = int(value)  # 更新 device_map 中的 value[0]

def update_config_value(key, new_value, filename="config.txt"):
    lines = []
    found = False

    with open(filename, "r") as f:
        for line in f:
            if line.startswith(f"{key}="):
                lines.append(f"{key}={new_value}\n")
                found = True
            else:
                lines.append(line)

    if not found:
        lines.append(f"{key}={new_value}\n")

    with open(filename, "w") as f:
        f.writelines(lines)



